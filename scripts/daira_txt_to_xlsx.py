#!/usr/bin/env python3
"""
Export daira folders (daira_accounts.txt + commune *.txt) to one .xlsx per daira.
Standalone: does not use or import w13_extract.py.

Install: pip install -r scripts/requirements-daira-excel.txt
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "Missing dependency: openpyxl. Install with:\n"
        "  pip install -r scripts/requirements-daira-excel.txt",
        file=sys.stderr,
    )
    sys.exit(1)

# Label (annexe / PO / etc.) and username, separated by ->, -->, --->, etc.
_ARROW_SPLIT = re.compile(r"\s*-+>\s*")
_ORDER_PREFIX = re.compile(r"^\d+(?:-\d+)?\s*-\s*")
_PO_LEFT = re.compile(r"(?i)^PO\b")

HEADERS = ("Daïra", "Commune", "Annexe", "Identifiant")
DAIRA_ACCOUNTS_NAME = "daira_accounts.txt"


def strip_order_prefix(name: str) -> str:
    """Strip leading '002 - ' or '005-2 - ' style prefix from folder or file stem."""
    s = name.strip()
    m = _ORDER_PREFIX.match(s)
    if not m:
        return s
    return s[m.end() :]


def split_arrow_line(line: str) -> tuple[str, str] | None:
    line = line.strip()
    if not line:
        return None
    parts = _ARROW_SPLIT.split(line, maxsplit=1)
    if len(parts) != 2:
        return None
    left, right = parts[0].strip(), parts[1].strip()
    if not right:
        return None
    return left, right


def daira_display_name(folder: Path) -> str:
    return strip_order_prefix(folder.name)


def commune_display_name(txt_path: Path) -> str:
    return strip_order_prefix(txt_path.stem)


def collect_rows_for_daira(daira_dir: Path) -> list[tuple[str, str, str, str]]:
    daira_name = daira_display_name(daira_dir)
    rows: list[tuple[str, str, str, str]] = []

    accounts = daira_dir / DAIRA_ACCOUNTS_NAME
    if accounts.is_file():
        for raw in accounts.read_text(encoding="utf-8", errors="replace").splitlines():
            parsed = split_arrow_line(raw)
            if not parsed:
                continue
            _, username = parsed
            rows.append((daira_name, "", "", username))

    commune_files = sorted(
        p
        for p in daira_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".txt"
        and p.name != DAIRA_ACCOUNTS_NAME
    )
    for path in commune_files:
        commune_name = commune_display_name(path)
        for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
            parsed = split_arrow_line(raw)
            if not parsed:
                continue
            left, username = parsed
            if _PO_LEFT.match(left):
                rows.append((daira_name, "", "", username))
            else:
                rows.append((daira_name, commune_name, left, username))

    return rows


def write_workbook(out_path: Path, rows: list[tuple[str, str, str, str]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comptes"
    ws.append(list(HEADERS))
    for r in rows:
        ws.append(list(r))
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28
    wb.save(out_path)


def _match_only(folder_name: str, only: str) -> bool:
    only = only.strip()
    if not only:
        return False
    return only == folder_name or only in folder_name


def main() -> None:
    repo = Path(__file__).resolve().parent.parent
    default_dairas = repo / "out_w13_all_ordered_col1" / "dairas"
    default_out = repo / "out_w13_all_ordered_col1" / "daira_excel"

    p = argparse.ArgumentParser(description="Export daira text folders to Excel workbooks.")
    p.add_argument(
        "--dairas-root",
        type=Path,
        default=default_dairas,
        help=f"Folder containing daira subfolders (default: {default_dairas})",
    )
    p.add_argument(
        "--out-dir",
        type=Path,
        default=default_out,
        help=f"Single output folder for all .xlsx files (default: {default_out})",
    )
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument(
        "--all",
        action="store_true",
        help="Export every immediate child folder of --dairas-root",
    )
    g.add_argument(
        "--only",
        metavar="SUBSTRING",
        help="Export only the daira folder whose name equals or contains this string",
    )
    args = p.parse_args()

    dairas_root: Path = args.dairas_root.resolve()
    out_dir: Path = args.out_dir.resolve()
    if not dairas_root.is_dir():
        print(f"Not a directory: {dairas_root}", file=sys.stderr)
        sys.exit(1)

    candidates = sorted(
        d for d in dairas_root.iterdir() if d.is_dir() and not d.name.startswith(".")
    )
    if args.only:
        selected = [d for d in candidates if _match_only(d.name, args.only)]
        if not selected:
            print(
                f"No daira folder under {dairas_root} matches --only {args.only!r}",
                file=sys.stderr,
            )
            sys.exit(1)
        if len(selected) > 1:
            print(
                "Ambiguous --only; matches more than one folder:\n"
                + "\n".join(f"  {d.name}" for d in selected),
                file=sys.stderr,
            )
            sys.exit(1)
        to_export = selected
    else:
        to_export = candidates

    out_dir.mkdir(parents=True, exist_ok=True)
    for daira_dir in to_export:
        rows = collect_rows_for_daira(daira_dir)
        out_name = f"{daira_dir.name}.xlsx"
        out_path = out_dir / out_name
        write_workbook(out_path, rows)
        print(f"Wrote {out_path} ({len(rows)} data rows)")


if __name__ == "__main__":
    main()
